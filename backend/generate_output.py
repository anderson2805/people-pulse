import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

from backend.linkedin import classify_li
from backend.tech_domain import classify_tech_llm
from backend.ssoc import classify_ssoc_llm
from backend.affiliation import extract_affiliate_llm
import streamlit as st


def update_output_excel(template_path, output_path =  'presentation/Output.xlsx'):
    # Read the template Excel file
    template_df = pd.read_excel(template_path, header=2)
    # drop the first row
    template_df = template_df.drop(0)
    with st.status("Pulsing through the web for enrichment of data...", expanded=True) as status:
        li_urls = template_df['LinkedIn URL'].dropna().tolist()
        
        st.write("Updating LinkedIn Context...")
        li_contexts = classify_li(li_urls)
        
        def get_linkedin_info(url):
            return li_contexts.get(url, "No LinkedIn data found")
        
        template_df['LinkedIn Context'] = template_df['LinkedIn URL'].apply(get_linkedin_info)
        
        # Read the output Excel file, specifying the header row
        output_df = pd.read_excel(output_path, header=2)  # header is on row 3 (0-based index 2)
        
        # Get the common columns between template and output
        common_columns = list(set(template_df.columns) & set(output_df.columns))
        
        # Start updating from row 5 (0-based index 4)
        start_row = 4
        
        # Convert template DataFrame to a list of dictionaries
        template_records = template_df.to_dict('records')
        # Update the output DataFrame
        for record in template_records:
            # Create a dictionary with only the common columns
            update_dict = {col: record[col] for col in common_columns if col in record}
            
            # Find matching rows in output_df and update them, starting from row 5
            for index in range(start_row, len(output_df)):
                if all(output_df.loc[index, col] == record[col] for col in common_columns if pd.notna(record[col])):
                    output_df.loc[index, common_columns] = pd.Series(update_dict)
                    break
            else:
                # If no matching row found, append a new row
                output_df = pd.concat([output_df, pd.DataFrame([update_dict])], ignore_index=True)
        
        # Load the existing workbook
        workbook = openpyxl.load_workbook(output_path)
        sheet = workbook.active
        
        # Update the cells with new data, starting from row 5
        for r_idx, row in enumerate(dataframe_to_rows(output_df, index=False, header=False), start=start_row+1):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                
        st.write("Applying LLM, classifying Technology domains on individuals...")
        tech_results = classify_tech_llm(template_df)
        # into worksheet with openpyxl, sheet name = "Technology Domain"
        # Check if the worksheet exists, if not, create it
        if "Technology Domain" not in workbook.sheetnames:
            sheet = workbook.create_sheet("Technology Domain")
        else:
            sheet = workbook["Technology Domain"]
        # Add headers
        headers = ["Name", "Main Topic", "Sub Topic", "Focus Area"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
            sheet.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
            
        # Write data
        for row, result in enumerate(tech_results, start=2):
            sheet.cell(row=row, column=1, value=result['name'])
            sheet.cell(row=row, column=2, value=result['main_topic'])
            sheet.cell(row=row, column=3, value=result['sub_topic'])
            sheet.cell(row=row, column=4, value=result['focus_area'])
        
        st.write("Applying LLM, classifying SSOC 2024 code on individuals...")
        ssoc_results = classify_ssoc_llm(template_df)
        # into worksheet with openpyxl, sheet name = "Technology Domain"
        # Check if the worksheet exists, if not, create it
        if "SSOC" not in workbook.sheetnames:
            sheet = workbook.create_sheet("SSOC")
        else:
            sheet = workbook["SSOC"]
        # Add headers
        headers = ["Name", "Code", "Title", "Task"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
            # Bold the first row
            sheet.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)

        # Write data
        row = 2  # Start from row 2, as row 1 is now for headers
        for name, occupations in ssoc_results.items():
            for occupation in occupations:
                sheet.cell(row=row, column=1, value=name)
                sheet.cell(row=row, column=2, value=occupation['code'])
                sheet.cell(row=row, column=3, value=occupation['title'])
                sheet.cell(row=row, column=4, value=occupation['task'])
                row += 1

        st.write("Applying LLM, extracting affiliations of individuals or companies...")
        affiliation_results = extract_affiliate_llm(template_df)
        if "Affiliation" not in workbook.sheetnames:
            sheet = workbook.create_sheet("Affiliation")
        else:
            sheet = workbook["Affiliation"]
            
        headers = ["Name", "Organization", "Status"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
            sheet.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
            
        row = 2
        for name, affiliations in affiliation_results.items():
            for affiliation in affiliations:
                sheet.cell(row=row, column=1, value=name)
                sheet.cell(row=row, column=2, value=affiliation['organization'])
                sheet.cell(row=row, column=3, value=affiliation['status'])
                row += 1
        # Save the updated workbook with new filename to avoid overwriting the original file, output_path = '../presentation/Output.xlsx'
        # output_path = 'updated_output.xlsx'
        # workbook.save(output_path)
        status.update(
        label="Pulsing Completed!", state="complete", expanded=False
        )
        print(f"Workbook has been updated with data from {template_path}, preserving original formatting.")
    return workbook

# Usage
if __name__ == "__main__":
    update_output_excel('presentation/Template.xlsx', 'presentation/Output.xlsx')